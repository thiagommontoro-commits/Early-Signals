# ==========================================================================
# MÓDULO: COMMODITY INTELLIGENCE  (Early Signals - LATAM)
# --------------------------------------------------------------------------
# Lê a planilha mensal de preços (data/precos_agricolas_latest.xlsx) e calcula:
#   Preço atual | Δ M/M | Δ A/A | Média 5 anos | Percentil histórico | Farol
#   + Preços regionais em US$ (Soja/Milho MT)
#   + TENDÊNCIA (Opção C): direção projetada = momentum 3m + sazonalidade,
#     100% derivada da própria planilha (nenhuma fonte externa).
#
# MAPA DE COLUNAS (benchmark internacional) validado contra AGOSTO/2026:
#   Soja=Soja col15 CBOT | Milho=Milho col21 CBOT | Café=Demais AGRÍCOLAS col6 ICE
#   Açúcar=Cana col6 ICE | Algodão=Algodão col8 ICE | Trigo=Trigo col11 US SRW
# Regionais US$/60kg: Soja MT=col11 | Milho MT=col10
# ==========================================================================

import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    openpyxl = None

MESES_PT = ['JAN', 'FEV', 'MAR', 'ABR', 'MAI', 'JUN',
            'JUL', 'AGO', 'SET', 'OUT', 'NOV', 'DEZ']

FONTE_PUBLICA = "CME Group & ICE Futures (US)"

COMMODITIES = {
    "soja":    {"aba": "Soja",             "coluna": 15, "unidade": "US$/bushel", "icone": "🌱",
                "nome": {"pt": "Soja (CBOT)", "en": "Soybean (CBOT)", "es": "Soja (CBOT)"},
                "regionais": [{"uf": "MT", "coluna": 11, "unidade": "US$/60kg"}]},
    "milho":   {"aba": "Milho",            "coluna": 21, "unidade": "US$/bushel", "icone": "🌽",
                "nome": {"pt": "Milho (CBOT)", "en": "Corn (CBOT)", "es": "Maíz (CBOT)"},
                "regionais": [{"uf": "MT", "coluna": 10, "unidade": "US$/60kg"}]},
    "cafe":    {"aba": "Demais AGRÍCOLAS", "coluna": 6,  "unidade": "¢/lb", "icone": "☕",
                "nome": {"pt": "Café Arábica (ICE NY)", "en": "Arabica Coffee (ICE NY)", "es": "Café Arábica (ICE NY)"},
                "regionais": []},
    "acucar":  {"aba": "Cana",             "coluna": 6,  "unidade": "¢/lb", "icone": "🍬",
                "nome": {"pt": "Açúcar (ICE NY)", "en": "Sugar (ICE NY)", "es": "Azúcar (ICE NY)"},
                "regionais": []},
    "algodao": {"aba": "Algodão",          "coluna": 8,  "unidade": "¢/lb", "icone": "🧵",
                "nome": {"pt": "Algodão (ICE NY)", "en": "Cotton (ICE NY)", "es": "Algodón (ICE NY)"},
                "regionais": []},
    "trigo":   {"aba": "Trigo",            "coluna": 11, "unidade": "US$/t", "icone": "🌾",
                "nome": {"pt": "Trigo (US SRW)", "en": "Wheat (US SRW)", "es": "Trigo (US SRW)"},
                "regionais": []},
}

TOLERANCIA_MESES = 1


def _extrair_serie(ws, col_idx, row_start=7, row_end=500):
    serie, ano_atual = [], None
    for r in range(row_start, row_end + 1):
        ano_cell = ws.cell(row=r, column=1).value
        mes_cell = ws.cell(row=r, column=2).value
        val = ws.cell(row=r, column=col_idx).value
        if ano_cell is not None:
            try:
                ano_atual = int(ano_cell)
            except (ValueError, TypeError):
                pass
        if mes_cell is None or ano_atual is None:
            continue
        mes_str = str(mes_cell).strip().upper()[:3]
        if mes_str not in MESES_PT:
            continue
        if isinstance(val, (int, float)):
            serie.append((ano_atual, MESES_PT.index(mes_str) + 1, float(val)))
    return serie


def _percentil(valor, historico):
    if not historico:
        return None
    return round(100.0 * sum(1 for v in historico if v <= valor) / len(historico), 0)


def _farol(delta_mm, valor_atual, media_5a):
    if delta_mm is None or media_5a is None:
        return "incerto"
    if delta_mm > 0.5 and valor_atual >= media_5a:
        return "positivo"
    if delta_mm < -0.5 and valor_atual < media_5a:
        return "negativo"
    return "incerto"


def _tendencia_futura(serie):
    """
    OPÇÃO C — Tendência direcional projetada, 100% da própria planilha.
    Momentum (60%) + Sazonalidade histórica (40%).
    Retorna dict compacto: {classe, seta, pt, en, es, forca}
    """
    if len(serie) < 4:
        return {"classe": "estavel", "seta": "▬",
                "pt": "Estável", "en": "Stable", "es": "Estable", "forca": 0.0}

    ult = [v for (_, _, v) in serie[-4:]]
    variacoes = []
    for i in range(1, len(ult)):
        if ult[i-1]:
            variacoes.append(100.0 * (ult[i] - ult[i-1]) / ult[i-1])
    momentum = sum(variacoes) / len(variacoes) if variacoes else 0.0

    _, mes_atual, _ = serie[-1]
    mes_seguinte = mes_atual % 12 + 1
    por_mes = {(ano, mes): val for (ano, mes, val) in serie}
    saz_variacoes = []
    for (ano, mes), val in por_mes.items():
        if mes == mes_atual:
            ano_seg = ano + 1 if mes_seguinte == 1 else ano
            prox = por_mes.get((ano_seg, mes_seguinte))
            if prox and val:
                saz_variacoes.append(100.0 * (prox - val) / val)
    sazonalidade = sum(saz_variacoes) / len(saz_variacoes) if saz_variacoes else 0.0

    score = 0.6 * momentum + 0.4 * sazonalidade

    if score > 1.0:
        return {"classe": "positivo", "seta": "▲",
                "pt": "Alta", "en": "Up", "es": "Alza", "forca": round(score, 1)}
    if score < -1.0:
        return {"classe": "negativo", "seta": "▼",
                "pt": "Baixa", "en": "Down", "es": "Baja", "forca": round(score, 1)}
    return {"classe": "estavel", "seta": "▬",
            "pt": "Estável", "en": "Stable", "es": "Estable", "forca": round(score, 1)}


def _calc_regionais(ws, cfg):
    saida = []
    for reg in cfg.get("regionais", []):
        serie = _extrair_serie(ws, reg["coluna"])
        if not serie:
            continue
        _, _, atual = serie[-1]
        delta_mm = None
        if len(serie) >= 2 and serie[-2][2]:
            delta_mm = round(100.0 * (atual - serie[-2][2]) / serie[-2][2], 1)
        saida.append({"uf": reg["uf"], "unidade": reg["unidade"],
                      "preco": round(atual, 2), "delta_mm": delta_mm})
    return saida


def calcular_commodity(serie, cfg, ws=None):
    if not serie:
        return None
    ano_a, mes_a, preco_atual = serie[-1]
    delta_mm = None
    if len(serie) >= 2 and serie[-2][2]:
        delta_mm = round(100.0 * (preco_atual - serie[-2][2]) / serie[-2][2], 1)
    delta_aa = None
    for (ano, mes, val) in serie:
        if (ano, mes) == (ano_a - 1, mes_a) and val:
            delta_aa = round(100.0 * (preco_atual - val) / val, 1)
            break
    ult_60 = [v for (_, _, v) in serie[-60:]]
    media_5a = round(sum(ult_60) / len(ult_60), 2) if ult_60 else None
    historico = [v for (_, _, v) in serie]
    return {
        "id": None, "nome": cfg["nome"], "icone": cfg["icone"], "unidade": cfg["unidade"],
        "preco_atual": round(preco_atual, 2), "ano_ref": ano_a, "mes_ref_idx": mes_a,
        "mes_ref": f"{MESES_PT[mes_a-1]}/{ano_a}",
        "delta_mm": delta_mm, "delta_aa": delta_aa, "media_5a": media_5a,
        "minimo": round(min(historico), 2), "maximo": round(max(historico), 2),
        "percentil": _percentil(preco_atual, historico),
        "tendencia": _farol(delta_mm, preco_atual, media_5a),
        "tendencia_futura": _tendencia_futura(serie),
        "regionais": _calc_regionais(ws, cfg) if ws is not None else [],
        "fonte": FONTE_PUBLICA,
    }


def _checar_frescor(resultados):
    if not resultados:
        return resultados, None
    hoje = datetime.datetime.now()
    ano_ref, mes_ref_idx = max((r["ano_ref"], r["mes_ref_idx"]) for r in resultados)
    dif = (hoje.year - ano_ref) * 12 + (hoje.month - mes_ref_idx)
    aviso = None
    if dif > TOLERANCIA_MESES:
        aviso = (f"⚠️ ATENÇÃO: planilha traz dados até {MESES_PT[mes_ref_idx-1]}/{ano_ref}, "
                 f"mas estamos em {MESES_PT[hoje.month-1]}/{hoje.year} ({dif} mês(es) de defasagem).")
        print(aviso)
    for r in resultados:
        r["dados_desatualizados"] = bool(aviso)
    return resultados, aviso


def processar_commodities(xlsx_path):
    xlsx_path = Path(xlsx_path)
    if openpyxl is None:
        print("❌ openpyxl não instalado."); return [], "erro_openpyxl"
    if not xlsx_path.exists():
        print(f"⚠️ Planilha não encontrada em '{xlsx_path}'."); return [], "arquivo_ausente"
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        print(f"❌ Erro ao abrir planilha: {e}"); return [], "erro_leitura"
    resultados = []
    for cid, cfg in COMMODITIES.items():
        try:
            if cfg["aba"] not in wb.sheetnames:
                print(f"   ⚠️ Aba '{cfg['aba']}' ausente ({cid})."); continue
            ws = wb[cfg["aba"]]
            serie = _extrair_serie(ws, cfg["coluna"])
            m = calcular_commodity(serie, cfg, ws=ws)
            if m:
                m["id"] = cid; resultados.append(m)
                tf = m["tendencia_futura"]
                print(f"   ✅ {cid}: {m['preco_atual']} {cfg['unidade']} ({m['mes_ref']}) "
                      f"M/M {m['delta_mm']}% | tendência: {tf['seta']} {tf['pt']} ({tf['forca']})")
        except Exception as e:
            print(f"   ❌ Erro em '{cid}': {e}")
    wb.close()
    return _checar_frescor(resultados)


if __name__ == "__main__":
    import json
    caminho = Path(__file__).parent / "data" / "precos_agricolas_latest.xlsx"
    dados, aviso = processar_commodities(caminho)
    print("\n=== AVISO ===", aviso or "✅ em dia")
    print(json.dumps(dados, ensure_ascii=False, indent=2))
